import openpyxl

wb = openpyxl.load_workbook('/home/homelinux/file_exchange/TEMPLATE_Timesheet_GoogleSheets.xlsx', data_only=False)
print("Sheet names:", wb.sheetnames)
if 'Summary' in wb.sheetnames:
    ws = wb['Summary']
    print("Max row:", ws.max_row, "Max col:", ws.max_column)
    for r in range(1, min(15, ws.max_row+1)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(10, ws.max_column+1))]
        print(f"Row {r}: {row_vals}")
