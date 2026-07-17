#!/usr/bin/env python3
"""
create_summary_sheet.py
-----------------------
Membuat tab 'Ringkasan' pada spreadsheet template dengan rumus-rumus
yang mereplikasi tampilan Data Preview pada UI web app.

Layout mengikuti Screenshot 1 (Capture):
┌──────────────────────────────────────────────────────┬────────────────────┐
│  Total Open Tiket │ Total Closed Tiket │ Hari Kerja │ Hari Off           │
│  [FORMULA]        │ [FORMULA]          │ [FORMULA]  │ [FORMULA]          │
├──────────────────────────────────────────────────────┼────────────────────┤
│                                                      │                    │
│  Tanggal│ Shift │ Jam    │ Open │ Closed │ Remark    │  Nama Petugas      │
│  (rows 1-31 with formulas)                           │  [DROPDOWN]        │
│                                                      │                    │
└──────────────────────────────────────────────────────┴────────────────────┘

Rumus:
- Shift: INDEX/MATCH dari 'Jadwal Shifting' berdasarkan nama terpilih
- Jam: SWITCH/IFS berdasarkan nilai shift (1→06:00-15:00, 2→14:00-23:00, dll)
- Open: COUNTIFS mencocokkan Nama Petugas + tanggal di 'Open Insiden'
- Closed: COUNTIFS mencocokkan Nama Petugas + tanggal di 'Closed Insiden'
- Remark: IFS berdasarkan nilai shift
- Summary: SUM dan COUNTIF dari kolom tabel

Kompatibel dengan Google Sheets (setelah import .xlsx).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import calendar
import os
import sys

# ── Path ──────────────────────────────────────────────────────────────
file_path = os.path.expanduser('~/file_exchange/TEMPLATE_Timesheet_GoogleSheets.xlsx')
wb = openpyxl.load_workbook(file_path)

# ── Deteksi bulan/tahun ───────────────────────────────────────────────
# CLI args: --month M --year Y (opsional)
TEMPLATE_YEAR = 2026
TEMPLATE_MONTH = 7  # Default Juli 2026

# Parse CLI arguments
cli_args = sys.argv[1:]
for i, arg in enumerate(cli_args):
    if arg == '--month' and i + 1 < len(cli_args):
        TEMPLATE_MONTH = int(cli_args[i + 1])
    elif arg == '--year' and i + 1 < len(cli_args):
        TEMPLATE_YEAR = int(cli_args[i + 1])

# Coba deteksi bulan dari data Open Insiden jika ada datetime data
if 'Open Insiden' in wb.sheetnames:
    ws_open = wb['Open Insiden']
    # Cek baris 2 (data pertama setelah header)
    first_dt = ws_open.cell(row=2, column=1).value
    if first_dt and hasattr(first_dt, 'month'):
        TEMPLATE_MONTH = first_dt.month
        TEMPLATE_YEAR = first_dt.year

num_days = calendar.monthrange(TEMPLATE_YEAR, TEMPLATE_MONTH)[1]
MONTH_NAMES = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]
month_label = f"{MONTH_NAMES[TEMPLATE_MONTH - 1]} {TEMPLATE_YEAR}"

# ── Deteksi nama karyawan dari Jadwal Shifting ────────────────────────
employee_names = []
if 'Jadwal Shifting' in wb.sheetnames:
    ws_sched = wb['Jadwal Shifting']
    import re
    for row_idx in range(3, ws_sched.max_row + 1):
        name = ws_sched.cell(row=row_idx, column=1).value
        if name:
            name_str = str(name).strip()
            if name_str and name_str.lower() != 'nan':
                if re.search(r'backup|izin sakit|tanggal\s+\d+', name_str, re.IGNORECASE):
                    continue
                employee_names.append(name_str)

# ── Hapus sheet lama jika ada, buat baru ──────────────────────────────
if 'Ringkasan' in wb.sheetnames:
    del wb['Ringkasan']

ws = wb.create_sheet('Ringkasan', 0)
wb.active = 0

# ══════════════════════════════════════════════════════════════════════
# STYLE DEFINITIONS
# ══════════════════════════════════════════════════════════════════════
FONT_FAMILY = 'Segoe UI'

# Summary header bar
summary_header_font = Font(name=FONT_FAMILY, size=10, bold=True, color='1E3A8A')
summary_header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
summary_header_align = Alignment(horizontal='center', vertical='center')

# Summary value
summary_value_font = Font(name=FONT_FAMILY, size=18, bold=True, color='1E3A8A')
summary_value_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
summary_value_align = Alignment(horizontal='center', vertical='center')

# Table header (dark blue)
table_header_font = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')
table_header_fill = PatternFill(start_color='1E3764', end_color='1E3764', fill_type='solid')
table_header_align = Alignment(horizontal='center', vertical='center')

# Nama Petugas header
nama_header_font = Font(name=FONT_FAMILY, size=10, bold=True, color='FFFFFF')
nama_header_fill = PatternFill(start_color='1E3764', end_color='1E3764', fill_type='solid')

# Table body
body_font = Font(name=FONT_FAMILY, size=10, color='333333')
body_font_center = Font(name=FONT_FAMILY, size=10, color='333333')
body_align_center = Alignment(horizontal='center', vertical='center')
body_align_left = Alignment(horizontal='left', vertical='center')

# Zebra striping
zebra_fill_1 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
zebra_fill_2 = PatternFill(start_color='E8ECF0', end_color='E8ECF0', fill_type='solid')

# OFF row fill
off_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

# Dropdown style
dropdown_font = Font(name=FONT_FAMILY, size=12, bold=True, color='1E3A8A')
dropdown_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
dropdown_border = Border(
    left=Side(style='medium', color='1E3764'),
    right=Side(style='medium', color='1E3764'),
    top=Side(style='medium', color='1E3764'),
    bottom=Side(style='medium', color='1E3764')
)

# Borders
thin_border = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)

no_border = Border()

# ══════════════════════════════════════════════════════════════════════
# LAYOUT - Mengikuti Screenshot 1
# ══════════════════════════════════════════════════════════════════════

# Jumlah baris employee di Jadwal Shifting (untuk MATCH range)
num_employees = len(employee_names)
sched_last_row = 2 + num_employees  # Row 3 sampai sched_last_row di Jadwal Shifting

# ── ROW 1: Summary Headers ───────────────────────────────────────────
# Kolom A-B: Total Open Tiket | C-D: Total Closed Tiket | E-F: Hari Kerja | G-H: Hari Off (sesuai screenshot)
# Kita gunakan merge cells per summary item
summary_labels = ['Total Open Tiket', 'Total Closed Tiket', 'Hari Kerja', 'Hari Off']
summary_col_starts = [1, 3, 5, 7]  # A, C, E, G (masing-masing merge 2 kolom)

for i, (label, col_start) in enumerate(zip(summary_labels, summary_col_starts)):
    col_end = col_start + 1
    ws.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
    cell = ws.cell(row=1, column=col_start, value=label)
    cell.font = summary_header_font
    cell.fill = summary_header_fill
    cell.alignment = summary_header_align
    # Style untuk merged cell partner
    cell2 = ws.cell(row=1, column=col_end)
    cell2.fill = summary_header_fill

ws.row_dimensions[1].height = 24

# ── ROW 2: Summary Values ────────────────────────────────────────────
# Rumus summary — referensi ke tabel di bawah
# Tabel data dimulai dari row 6 (header) → data row 7 sampai 7+num_days-1
data_start_row = 7
data_end_row = data_start_row + num_days - 1

# Total Open Tiket = SUM kolom D (Open) dari tabel
# Total Closed Tiket = SUM kolom E (Closed) dari tabel
# Hari Kerja = jumlah hari yang shift-nya bukan OFF dan bukan IS
# Hari Off = jumlah hari yang shift-nya OFF

summary_formulas = [
    f'=SUM(D{data_start_row}:D{data_end_row})',                                                    # Total Open Tiket
    f'=SUM(E{data_start_row}:E{data_end_row})',                                                    # Total Closed Tiket
    f'=COUNTIF(B{data_start_row}:B{data_end_row},"<>OFF")-COUNTIF(B{data_start_row}:B{data_end_row},"IS")-COUNTBLANK(B{data_start_row}:B{data_end_row})',  # Hari Kerja
    f'=COUNTIF(B{data_start_row}:B{data_end_row},"OFF")',                                          # Hari Off
]

for i, (formula, col_start) in enumerate(zip(summary_formulas, summary_col_starts)):
    col_end = col_start + 1
    ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
    cell = ws.cell(row=2, column=col_start, value=formula)
    cell.font = summary_value_font
    cell.fill = summary_value_fill
    cell.alignment = summary_value_align
    cell.border = thin_border
    cell2 = ws.cell(row=2, column=col_end)
    cell2.fill = summary_value_fill

ws.row_dimensions[2].height = 36

# ── ROW 3-4: Spacer ──────────────────────────────────────────────────
ws.row_dimensions[3].height = 6
ws.row_dimensions[4].height = 6

# ── ROW 5: Blank / Spacer ────────────────────────────────────────────
ws.row_dimensions[5].height = 6

# ── ROW 6: Table Headers + Nama Petugas Header ───────────────────────
table_headers = ['Tanggal', 'Shift', 'Jam', 'Open', 'Closed', 'Remark']
for col_idx, header in enumerate(table_headers, start=1):
    cell = ws.cell(row=6, column=col_idx, value=header)
    cell.font = table_header_font
    cell.fill = table_header_fill
    cell.alignment = table_header_align

# Nama Petugas header di kolom H (kolom 8)
cell_nama = ws.cell(row=6, column=8, value='Nama Petugas')
cell_nama.font = nama_header_font
cell_nama.fill = nama_header_fill
cell_nama.alignment = Alignment(horizontal='left', vertical='center')
# Merge H6:I6 untuk header yang lebih lebar
ws.merge_cells(start_row=6, start_column=8, end_row=6, end_column=9)

ws.row_dimensions[6].height = 28

# ── ROW 7: Nama Petugas Value (Dropdown) ──────────────────────────────
# Dropdown di H7 yang referensi nama dari Jadwal Shifting
cell_nama_val = ws.cell(row=7, column=8)
cell_nama_val.font = dropdown_font
cell_nama_val.fill = dropdown_fill
cell_nama_val.border = dropdown_border
cell_nama_val.alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells(start_row=7, start_column=8, end_row=7, end_column=9)

# Set default value (nama pertama dari daftar)
if employee_names:
    cell_nama_val.value = employee_names[0]

# Data Validation: dropdown dari nama karyawan di Jadwal Shifting
# Gunakan referensi langsung ke range di Jadwal Shifting
dv_formula = f"'Jadwal Shifting'!$A$3:$A${sched_last_row}"
dv = DataValidation(
    type="list",
    formula1=dv_formula,
    allow_blank=False
)
dv.error = 'Pilih nama dari daftar dropdown'
dv.errorTitle = 'Nama tidak valid'
dv.prompt = 'Klik untuk memilih nama konsultan'
dv.promptTitle = 'Pilih Konsultan'
ws.add_data_validation(dv)
dv.add('H7')

# ── ROWS 7 to 7+num_days-1: Data Rows with Formulas ──────────────────
# Helper: mendapatkan kolom letter di Jadwal Shifting untuk hari ke-N
# Jadwal Shifting: col A=nama, col B=hari 1, col C=hari 2, ... col (day+1)=hari day

for day in range(1, num_days + 1):
    row_num = data_start_row + day - 1  # data_start_row = 7
    sched_col_letter = get_column_letter(day + 1)  # day 1 → col B, day 2 → col C, etc.

    # A: Tanggal (DD/MM/YYYY)
    date_str = f"{day:02d}/{TEMPLATE_MONTH:02d}/{TEMPLATE_YEAR}"
    cell_date = ws.cell(row=row_num, column=1, value=date_str)

    # B: Shift — INDEX/MATCH dari Jadwal Shifting berdasarkan nama di H7
    shift_formula = (
        f"=IFERROR("
        f"INDEX('Jadwal Shifting'!{sched_col_letter}$3:{sched_col_letter}${sched_last_row},"
        f"MATCH($H$7,'Jadwal Shifting'!$A$3:$A${sched_last_row},0))"
        f",\"\")"
    )
    ws.cell(row=row_num, column=2, value=shift_formula)

    # C: Jam — berdasarkan nilai shift dari config.py
    # Shift 1: 06:00 — 15:00
    # Shift 2: 14:00 — 23:00
    # Shift 3: 22:00 — 07:00
    # Shift 1.2: 06:00 — 23:00
    # Shift 2.3: 14:00 — 07:00
    # OFF/IS: kosong
    jam_formula = (
        f'=IF(B{row_num}="1","06:00 — 15:00",'
        f'IF(B{row_num}="2","14:00 — 23:00",'
        f'IF(B{row_num}="3","22:00 — 07:00",'
        f'IF(B{row_num}="1.2","06:00 — 23:00",'
        f'IF(B{row_num}="2.3","14:00 — 07:00",'
        f'"")))))' 
    )
    ws.cell(row=row_num, column=3, value=jam_formula)

    # D: Open Ticket — COUNTIFS mencocokkan Nama Petugas + tanggal hari ini
    # 'Open Insiden' kolom A = datetime, kolom C = Nama Petugas
    # Kita perlu mencocokkan tanggal (hanya tanggalnya, bukan waktu)
    # Gunakan COUNTIFS dengan range tanggal: >= awal hari AND < awal hari berikutnya
    # Karena di Google Sheets datetime disimpan sebagai angka:
    #   >= DATE(year,month,day) AND < DATE(year,month,day+1)

    if day < num_days:
        next_day = day + 1
        next_month = TEMPLATE_MONTH
        next_year = TEMPLATE_YEAR
    else:
        # Hari terakhir bulan → next date = tanggal 1 bulan berikutnya
        next_day = 1
        if TEMPLATE_MONTH == 12:
            next_month = 1
            next_year = TEMPLATE_YEAR + 1
        else:
            next_month = TEMPLATE_MONTH + 1
            next_year = TEMPLATE_YEAR

    open_formula = (
        f'=IF(B{row_num}="OFF","",'
        f'IF(B{row_num}="IS","",'
        f'IF(B{row_num}="",-1,'  # -1 sebagai marker row kosong (bulan < 31 hari)
        f'COUNTIFS('
        f"'Open Insiden'!$C:$C,$H$7,"
        f"'Open Insiden'!$A:$A,\">=\"&DATE({TEMPLATE_YEAR},{TEMPLATE_MONTH},{day}),"
        f"'Open Insiden'!$A:$A,\"<\"&DATE({next_year},{next_month},{next_day})"
        f'))))'
    )
    ws.cell(row=row_num, column=4, value=open_formula)

    closed_formula = (
        f'=IF(B{row_num}="OFF","",'
        f'IF(B{row_num}="IS","",'
        f'IF(B{row_num}="",-1,'
        f'COUNTIFS('
        f"'Closed Insiden'!$A:$A,\">=\"&DATE({TEMPLATE_YEAR},{TEMPLATE_MONTH},{day}),"
        f"'Closed Insiden'!$A:$A,\"<\"&DATE({next_year},{next_month},{next_day}),"
        f"'Closed Insiden'!$C:$C,$H$7"
        f'))))'
    )
    ws.cell(row=row_num, column=5, value=closed_formula)

    # F: Remark — berdasarkan shift
    remark_formula = (
        f'=IF(B{row_num}="OFF","OFF",'
        f'IF(B{row_num}="IS","Izin Sakit",'
        f'IF(B{row_num}="1","Shift 1",'
        f'IF(B{row_num}="2","Shift 2",'
        f'IF(B{row_num}="3","Shift 3",'
        f'IF(B{row_num}="1.2","Shift 1&2",'
        f'IF(B{row_num}="2.3","Shift 2&3",'
        f'"")))))))' 
    )
    ws.cell(row=row_num, column=6, value=remark_formula)

    # ── Styling ─────────────────────────────────
    row_fill = zebra_fill_2 if day % 2 == 0 else zebra_fill_1
    for col in range(1, 7):
        cell = ws.cell(row=row_num, column=col)
        cell.font = body_font
        cell.fill = row_fill
        cell.border = thin_border
        if col in [1, 2, 4, 5]:
            cell.alignment = body_align_center
        elif col == 3:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = body_align_left

# ── Column Widths ─────────────────────────────────────────────────────
col_widths = {
    1: 14,   # A: Tanggal
    2: 10,   # B: Shift
    3: 18,   # C: Jam
    4: 12,   # D: Open
    5: 12,   # E: Closed
    6: 16,   # F: Remark
    7: 3,    # G: spacer
    8: 22,   # H: Nama Petugas
    9: 10,   # I: extended merge
}
for col_idx, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Freeze panes (freeze header row) ─────────────────────────────────
ws.freeze_panes = 'A7'

# ── Print settings ────────────────────────────────────────────────────
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

# ══════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════
wb.save(file_path)
print(f"✅ Tab 'Ringkasan' berhasil dibuat di: {file_path}")
print(f"   Bulan: {month_label} ({num_days} hari)")
print(f"   Jumlah karyawan terdeteksi: {len(employee_names)}")
print(f"   Nama karyawan: {', '.join(employee_names)}")
print(f"   Dropdown referensi: 'Jadwal Shifting'!$A$3:$A${sched_last_row}")
print(f"   Data rows: {data_start_row} - {data_end_row}")
print()
print("📋 Instruksi selanjutnya:")
print("   1. Buka file di Google Sheets (import atau upload)")
print("   2. Tab 'Ringkasan' sudah aktif di posisi pertama")
print("   3. Pilih nama konsultan dari dropdown di kolom 'Nama Petugas' (H7)")
print("   4. Semua data (Shift, Jam, Open, Closed, Remark) akan otomatis berubah")
print("   5. Summary cards (Total Open/Closed Tiket, Hari Kerja, Hari Off) juga otomatis update")
