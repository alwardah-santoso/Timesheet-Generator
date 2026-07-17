import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

file_path = '/home/homelinux/file_exchange/TEMPLATE_Timesheet_GoogleSheets.xlsx'
wb = openpyxl.load_workbook(file_path)

if 'Summary' in wb.sheetnames:
    ws = wb['Summary']
    ws.delete_rows(1, ws.max_row)
else:
    ws = wb.create_sheet('Summary', 0)  # place as first tab

def col_letter(n):
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

# Styles
title_font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='334155', end_color='334155', fill_type='solid')

bold_font = Font(name='Segoe UI', size=11, bold=True)
regular_font = Font(name='Segoe UI', size=11)

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# 1. Title
ws.append(["TIMESHEET CLEANER — SUMMARY & DATA PREVIEW", "", "", "", "", ""])
ws.merge_cells("A1:F1")
for c in range(1, 7):
    cell = ws.cell(row=1, column=c)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

ws.append([])

# 2. Consultant Selection
ws.append(["Pilih Nama Konsultan:", "Ulil Amri", "", "Periode:", "Juni 2026", ""])
ws.cell(row=3, column=1).font = bold_font
ws.cell(row=3, column=2).font = bold_font
ws.cell(row=3, column=2).fill = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid') # Yellow accent
ws.cell(row=3, column=4).font = bold_font
ws.cell(row=3, column=5).font = bold_font

ws.append([])

# 3. Monthly Statistics Section
ws.append(["STATISTIK BULANAN", "", "", "", "", ""])
ws.cell(row=5, column=1).font = Font(name='Segoe UI', size=12, bold=True, color='1E40AF')

ws.append([
    "Hari Kerja", '=COUNTIF(B11:B40, "<>OFF")-COUNTIF(B11:B40, "IS")',
    "Hari OFF", '=COUNTIF(B11:B40, "OFF")',
    "Open Ticket", '=SUM(D11:D40)',
    "Closed Ticket", '=SUM(E11:E40)'
])
for col_idx in [1, 3, 5]:
    ws.cell(row=6, column=col_idx).font = bold_font
for col_idx in [2, 4, 6]:
    ws.cell(row=6, column=col_idx).font = bold_font
    ws.cell(row=6, column=col_idx).alignment = Alignment(horizontal='center')

ws.append([])
ws.append(["DATA PREVIEW HARIAN", "", "", "", "", ""])
ws.cell(row=8, column=1).font = Font(name='Segoe UI', size=12, bold=True, color='1E40AF')

# 4. Preview Table Headers
headers = ["Tanggal", "Shift", "Jam Kerja", "Open Ticket", "Closed Ticket", "Remark"]
ws.append(headers)
for c in range(1, 7):
    cell = ws.cell(row=9, column=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[9].height = 24

# 5. Daily Rows 1 to 30
for day in range(1, 31):
    row_num = 9 + day
    c_letter = col_letter(day + 1)
    date_str = f"{day:02d}/06/2026"
    
    shift_formula = f"=INDEX('Jadwal Shifting'!{c_letter}$3:{c_letter}$17, MATCH($B$3, 'Jadwal Shifting'!$A$3:$A$17, 0))"
    jam_formula = f'=IFS(B{row_num}="S1", "07:30 - 16:30", B{row_num}="S2", "15:30 - 00:30", B{row_num}="S3", "23:30 - 08:30", B{row_num}="S12", "07:30 - 00:30", B{row_num}="S23", "15:30 - 08:30", B{row_num}="OFF", "-", B{row_num}="IS", "-", TRUE, "-")'
    open_formula = f"=COUNTIFS('Open Insiden'!$C:$C, $B$3, 'Open Insiden'!$A:$A, A{row_num}&\"*\")"
    closed_formula = f"=COUNTIFS('Closed Insiden'!$C:$C, $B$3, 'Closed Insiden'!$A:$A, A{row_num}&\"*\")"
    remark_formula = f'=IFS(B{row_num}="OFF", "OFF", B{row_num}="IS", "Izin Sakit", B{row_num}="S1", "Shift 1 (Pagi)", B{row_num}="S2", "Shift 2 (Sore)", B{row_num}="S3", "Shift 3 (Malam)", B{row_num}="S12", "Shift 1 & 2", B{row_num}="S23", "Shift 2 & 3", TRUE, B{row_num})'

    ws.append([date_str, shift_formula, jam_formula, open_formula, closed_formula, remark_formula])
    for c in range(1, 7):
        cell = ws.cell(row=row_num, column=c)
        cell.font = regular_font
        cell.border = thin_border
        if c in [1, 2, 4, 5]:
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(horizontal='left')
    ws.row_dimensions[row_num].height = 20

# Column widths
widths = {1: 15, 2: 12, 3: 20, 4: 15, 5: 15, 6: 25}
for col_idx, width in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

wb.save(file_path)
print("Saved TEMPLATE_Timesheet_GoogleSheets.xlsx successfully!")
