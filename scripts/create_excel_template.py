#!/usr/bin/env python3
"""
Create a clean Excel template workbook for Google Sheets import.

Template ini disesuaikan dengan format yang dibaca oleh
timesheet-cleaner-fixed/core.py agar 100% kompatibel.

Format yang harus cocok:
  - Jadwal Shifting: Row 1 = day numbers (int), Row 2 = "NAMA" + day-of-week,
    Row 3+ = nama karyawan + shift values
  - Open Insiden: Header "Detected Time / From" | "Nama Insiden" | "Nama Petugas"
  - Closed Insiden: Header "ClosedDateTime" | "Nama Insiden" | "Nama Petugas"
  - Backup: Header "Tanggal" | "Nama Karyawan" | "Nama Backup" | "Shift"
  - Notes: Tanpa header, tanpa prefix "- " (core.py strips prefix sendiri)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import calendar
from datetime import datetime

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
font_body = Font(name="Arial", size=10)
font_bold = Font(name="Arial", size=10, bold=True)
font_subheader = Font(name="Arial", size=9, bold=True, color="666666")
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Warna sel yang dikenali oleh core.py (lihat core.py baris 284-298):
#   FFFFFF00 → yellow (Notes baris 1-4)
#   FF00FF00 → green  (Notes baris 5-7)
#   FF1010EA → blue   (OFF)
#   FFFFFFFF → white  (semua notes)
#   theme    → brown  (semua notes)
fill_yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
fill_green = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
fill_blue = PatternFill(start_color="FF1010EA", end_color="FF1010EA", fill_type="solid")
fill_off = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# ── Bulan/tahun template (bisa diubah) ───────────────────
TEMPLATE_YEAR = 2026
TEMPLATE_MONTH = 7  # Juli
num_days = calendar.monthrange(TEMPLATE_YEAR, TEMPLATE_MONTH)[1]

# Hitung hari dalam minggu untuk label day-of-week
DAY_ABBR = ['S', 'S', 'R', 'K', 'J', 'S', 'M']  # Sen=S, Sel=S, Rab=R, Kam=K, Jum=J, Sab=S, Min=M
day_labels = []
for d in range(1, num_days + 1):
    dow = datetime(TEMPLATE_YEAR, TEMPLATE_MONTH, d).weekday()  # 0=Monday
    day_labels.append(DAY_ABBR[dow])

# ──────────────────────────────────────────────────────────
# TAB 1: Jadwal Shifting
# ──────────────────────────────────────────────────────────
# Format yang dibaca core.py:
#   Row 1 (index 0): kosong | 1 | 2 | 3 | ... | num_days  (integer day numbers)
#   Row 2 (index 1): "NAMA" | day-of-week abbreviations   (dilewati oleh core.py)
#   Row 3+ (index 2+): nama karyawan | shift values per hari
#
# core.py baris 266-271 mendeteksi num_days dengan int(val) pada row index 0
# core.py baris 239-244 membaca nama dari row index 2+
# ──────────────────────────────────────────────────────────
ws1 = wb.create_sheet(title="Jadwal Shifting")

# Row 1: Kosong + day numbers (sebagai integer agar core.py bisa int(val))
ws1.cell(row=1, column=1, value=None)
for day in range(1, num_days + 1):
    c = ws1.cell(row=1, column=day + 1, value=day)  # integer, bukan string
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

# Row 2: "NAMA" + day-of-week abbreviations (dilewati oleh core.py, hanya visual)
ws1.cell(row=2, column=1, value="NAMA").font = font_subheader
for day_idx, label in enumerate(day_labels):
    c = ws1.cell(row=2, column=day_idx + 2, value=label)
    c.font = font_subheader
    c.alignment = align_center

# Sample data (Row 3+)
# Shift: 1, 2, 3, 1.2, 2.3, OFF, IS* (lihat core.py baris 109-175)
sample_sched = [
    ("Achmad Rizki Santoso",    ["1","2","3","OFF","1.2","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1"]),
    ("Fajaril Iqbal Alrasyid",  ["2","3","OFF","1","2","3","OFF","1.2","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1","2"]),
    ("Ihwan Kurniawan",         ["OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1.2","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF"]),
    ("Ulil Amri",               ["3","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3","OFF","1.2","OFF","1","2","3","OFF","1","2","3","OFF","1","2","3"]),
]

for r_idx, (name, shifts) in enumerate(sample_sched, start=3):
    c_name = ws1.cell(row=r_idx, column=1, value=name)
    c_name.font = font_bold
    c_name.alignment = align_left
    c_name.border = thin_border
    # Pad atau trim shifts agar sesuai num_days
    for d_idx in range(num_days):
        shift_val = shifts[d_idx] if d_idx < len(shifts) else "OFF"
        c = ws1.cell(row=r_idx, column=d_idx + 2, value=shift_val)
        c.font = font_body
        c.alignment = align_center
        c.border = thin_border
        if shift_val == "OFF":
            c.fill = fill_off
        elif shift_val in ["1.2", "2.3"]:
            c.font = Font(name="Arial", size=10, bold=True, color="006600")

# Demonstrasi warna sel (core.py membaca warna untuk filter notes):
# - Yellow (FFFFFF00): Notes baris 1-4
# - Green (FF00FF00): Notes baris 5-7
# - Blue (FF1010EA): OFF
# - White/Brown: Semua notes (1-7)
ws1.cell(row=3, column=2).fill = fill_yellow   # Hari 1 - contoh kuning
ws1.cell(row=3, column=3).fill = fill_green     # Hari 2 - contoh hijau
# OFF days bisa diberi warna biru
ws1.cell(row=3, column=5).fill = fill_blue      # Hari 4 (OFF) - contoh biru

ws1.column_dimensions['A'].width = 30
for col in range(2, num_days + 2):
    ws1.column_dimensions[get_column_letter(col)].width = 5.5

# ──────────────────────────────────────────────────────────
# TAB 2: Open Insiden
# ──────────────────────────────────────────────────────────
# Format yang dibaca core.py baris 346-361:
#   3 kolom: header mengandung "nama petugas" → mode multi-person
#   Kolom datetime: header mengandung 'time' atau 'detected'
#   Kolom insiden: header mengandung 'insiden' atau 'incident'
# ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet(title="Open Insiden")

headers_open = ["Detected Time / From", "Nama Insiden", "Nama Petugas"]
for col_idx, h in enumerate(headers_open, start=1):
    c = ws2.cell(row=1, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

sample_open = [
    ("2026-07-01 06:15:00", "UNCONFIRM",          "Achmad Rizki Santoso"),
    ("2026-07-01 08:30:00", "DISMANTLE",           "Achmad Rizki Santoso"),
    ("2026-07-01 14:20:00", "POWER ELECTRICITY",   "Fajaril Iqbal Alrasyid"),
    ("2026-07-02 10:00:00", "26J41246",            "Ihwan Kurniawan"),
    ("2026-07-02 15:45:00", "REAKTIVASI SIMCARD",  "Ulil Amri"),
]
for r_idx, row_val in enumerate(sample_open, start=2):
    for c_idx, val in enumerate(row_val, start=1):
        c = ws2.cell(row=r_idx, column=c_idx, value=val)
        c.font = font_body
        c.border = thin_border
        c.alignment = align_center if c_idx == 1 else align_left

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 30

# ──────────────────────────────────────────────────────────
# TAB 3: Closed Insiden
# ──────────────────────────────────────────────────────────
# Format yang dibaca core.py baris 370-391:
#   3 kolom: sama seperti Open Insiden
#   Header kolom 1: "ClosedDateTime"
# ──────────────────────────────────────────────────────────
ws3 = wb.create_sheet(title="Closed Insiden")

headers_closed = ["ClosedDateTime", "Nama Insiden", "Nama Petugas"]
for col_idx, h in enumerate(headers_closed, start=1):
    c = ws3.cell(row=1, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

sample_closed = [
    ("2026-07-01 07:00:00", "UNCONFIRM",          "Achmad Rizki Santoso"),
    ("2026-07-01 09:15:00", "DISMANTLE",           "Achmad Rizki Santoso"),
    ("2026-07-01 16:00:00", "POWER ELECTRICITY",   "Fajaril Iqbal Alrasyid"),
    ("2026-07-02 11:30:00", "REQUEST RANGING",     "Ihwan Kurniawan"),
    ("2026-07-02 17:00:00", "REAKTIVASI SIMCARD",  "Ulil Amri"),
]
for r_idx, row_val in enumerate(sample_closed, start=2):
    for c_idx, val in enumerate(row_val, start=1):
        c = ws3.cell(row=r_idx, column=c_idx, value=val)
        c.font = font_body
        c.border = thin_border
        c.alignment = align_center if c_idx == 1 else align_left

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 30

# ──────────────────────────────────────────────────────────
# TAB 4: Backup
# ──────────────────────────────────────────────────────────
# Format yang dibaca core.py baris 307-319:
#   Row 1: Header (dilewati, loop dari row index 1)
#   Col 0: Tanggal (int)
#   Col 1: Nama Karyawan (cocok dengan target_name)
#   Col 2: Nama Backup
#   Col 3: Shift (opsional)
# ──────────────────────────────────────────────────────────
ws4 = wb.create_sheet(title="Backup")

headers_backup = ["Tanggal", "Nama Karyawan", "Nama Backup", "Shift"]
for col_idx, h in enumerate(headers_backup, start=1):
    c = ws4.cell(row=1, column=col_idx, value=h)
    c.font = font_header
    c.fill = fill_header
    c.alignment = align_center

sample_backup = [
    (5,  "Achmad Rizki Santoso",   "Fajaril Iqbal Alrasyid", "1.2"),
    (8,  "Fajaril Iqbal Alrasyid", "Ihwan Kurniawan",        "1.2"),
    (14, "Ihwan Kurniawan",        "Ulil Amri",              "1.2"),
    (19, "Ulil Amri",              "Achmad Rizki Santoso",   "1.2"),
]
for r_idx, row_val in enumerate(sample_backup, start=2):
    for c_idx, val in enumerate(row_val, start=1):
        c = ws4.cell(row=r_idx, column=c_idx, value=val)
        c.font = font_body
        c.border = thin_border
        c.alignment = align_center if c_idx in [1, 4] else align_left

ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 12

# ──────────────────────────────────────────────────────────
# TAB 5: Notes
# ──────────────────────────────────────────────────────────
# Format yang dibaca core.py baris 336-343 + 403-411:
#   TANPA header (langsung isi dari Row 1)
#   TANPA prefix "- " (core.py baris 410 strips "- " sendiri)
#   Baris 1-4 → untuk warna kuning (yellow)
#   Baris 5-7 → untuk warna hijau (green)
#   Baris 1-7 → untuk warna putih/coklat (white/brown)
# ──────────────────────────────────────────────────────────
ws5 = wb.create_sheet(title="Notes")

sample_notes = [
    # Baris 1-4: Ditampilkan untuk sel berwarna KUNING
    "Monitoring WAG dan Telegram untuk insiden yang masuk",
    "Koordinasi dengan tim RCE terkait kendala perangkat",
    "Koordinasi dengan petugas IT lapangan dan vendor",
    "Membuat laporan pergantian shift dan serah terima tugas",
    # Baris 5-7: Ditampilkan untuk sel berwarna HIJAU
    "Pengecekan utilisasi link dan kestabilan koneksi",
    "Menindaklanjuti eskalasi tiket dari pelanggan atau operator",
    "Pencadangan (backup) log sistem sebelum akhir bulan",
]
for r_idx, note in enumerate(sample_notes, start=1):
    c = ws5.cell(row=r_idx, column=1, value=note)
    c.font = font_body
    c.border = thin_border
    c.alignment = align_left

ws5.column_dimensions['A'].width = 80

# ── Save ─────────────────────────────────────────────────
out_path = os.path.expanduser("~/file_exchange/TEMPLATE_Timesheet_GoogleSheets.xlsx")
wb.save(out_path)
print(f"✅ Template Excel saved to: {out_path}")
print(f"   Bulan: {TEMPLATE_MONTH}/{TEMPLATE_YEAR} ({num_days} hari)")
print(f"   Tabs: {', '.join(wb.sheetnames)}")
