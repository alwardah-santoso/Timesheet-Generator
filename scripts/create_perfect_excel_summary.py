import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

file_path = '/home/homelinux/file_exchange/TEMPLATE_Timesheet_GoogleSheets.xlsx'
wb = openpyxl.load_workbook(file_path)

# Ensure we ONLY modify or create the Summary sheet
if 'Summary' in wb.sheetnames:
    ws = wb['Summary']
    ws.delete_rows(1, ws.max_row)
else:
    ws = wb.create_sheet('Summary', 0)

# Make Summary the active/first sheet
wb.active = wb.sheetnames.index('Summary')

def col_letter(n):
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

# ── Color Palette & Styles ──────────────────────────────────────────
font_family = 'Segoe UI'

title_font = Font(name=font_family, size=15, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

section_font = Font(name=font_family, size=13, bold=True, color='1E3A8A')

label_font = Font(name=font_family, size=11, bold=True, color='334155')
dropdown_font = Font(name=font_family, size=12, bold=True, color='1E3A8A')
dropdown_fill = PatternFill(start_color='FEF08A', end_color='FEF08A', fill_type='solid') # soft yellow highlight

header_font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='312E81', end_color='312E81', fill_type='solid')

zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)
medium_border = Border(
    left=Side(style='medium', color='3b82f6'),
    right=Side(style='medium', color='3b82f6'),
    top=Side(style='medium', color='3b82f6'),
    bottom=Side(style='medium', color='3b82f6')
)

# ── 1. Title Banner (Row 1) ──────────────────────────────────────────
ws.append(["TIMESHEET CLEANER — SUMMARY & DATA PREVIEW", "", "", "", "", ""])
ws.merge_cells("A1:F1")
for col in range(1, 7):
    cell = ws.cell(row=1, column=col)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 36

# ── 2. Consultant Selection Dropdown (Row 2) ─────────────────────────
ws.append(["Pilih Nama Konsultan:", "Ulil Amri", "Periode:", "Juni 2026", "", ""])
ws.row_dimensions[2].height = 28

ws.cell(row=2, column=1).font = label_font
ws.cell(row=2, column=1).alignment = Alignment(horizontal='right', vertical='center')

cell_b2 = ws.cell(row=2, column=2)
cell_b2.font = dropdown_font
cell_b2.fill = dropdown_fill
cell_b2.border = medium_border
cell_b2.alignment = Alignment(horizontal='center', vertical='center')

ws.cell(row=2, column=3).font = label_font
ws.cell(row=2, column=3).alignment = Alignment(horizontal='right', vertical='center')

cell_d2 = ws.cell(row=2, column=4)
cell_d2.font = Font(name=font_family, size=12, bold=True, color='0F172A')
cell_d2.alignment = Alignment(horizontal='left', vertical='center')

# Add Data Validation Dropdown to B2
consultant_names = (
    "Anggit Puji Liestary,Nur Rifda Ellysa,Karina Ari Mukti,Sheren Alvionita Siahaan,"
    "Ihwan Kurniawan,Riyan Saputra,Fajaril Iqbal Alrasyid,Gilang Donny Karunia,Ulil Amri,"
    "Arif Pebrianto,Achmad Rizki Santoso,Juliansyah Ortega,Iman Arya Wisetiaputra,"
    "Muhamad Rizky Saputra,Abadi A.H Pangaribuan"
)
dv = DataValidation(type="list", formula1=f'"{consultant_names}"', allow_blank=True)
dv.error ='Nama konsultan tidak valid'
dv.errorTitle = 'Pilih dari daftar dropdown'
dv.prompt = 'Klik panah dropdown untuk memilih nama'
dv.promptTitle = 'Pilih Konsultan'
ws.add_data_validation(dv)
dv.add("B2")

ws.append([]) # Row 3 blank

# ── 3. Statistik Bulanan Section (Rows 4 & 5) ────────────────────────
ws.append(["STATISTIK BULANAN", "", "", "", "", ""])
ws.cell(row=4, column=1).font = section_font
ws.row_dimensions[4].height = 24

ws.append([
    "Hari Kerja", '=COUNTIF(B9:B38, "<>OFF")-COUNTIF(B9:B38, "IS")',
    "Hari OFF", '=COUNTIF(B9:B38, "OFF")',
    "Open Ticket", '=SUM(D9:D38)',
    "Closed Ticket", '=SUM(E9:E38)'
])
ws.row_dimensions[5].height = 26

stat_styles = [
    (1, 2, 'DBEAFE', '1E40AF'), # Hari Kerja (Blue)
    (3, 4, 'FEF9C3', '854D0E'), # Hari OFF (Yellow)
    (5, 6, 'DCFCE7', '166534'), # Open Ticket (Green)
]
for lbl_col, val_col, fill_color, text_color in stat_styles:
    for c in [lbl_col, val_col]:
        cell = ws.cell(row=5, column=c)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        cell.font = Font(name=font_family, size=11, bold=True, color=text_color)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

ws.append([]) # Row 6 blank

# ── 4. Tabel Preview Header (Rows 7 & 8) ─────────────────────────────
ws.append(["DATA PREVIEW HARIAN", "", "", "", "", ""])
ws.cell(row=7, column=1).font = section_font
ws.row_dimensions[7].height = 24

headers = ["Tanggal", "Shift", "Jam Kerja", "Open Ticket", "Closed Ticket", "Remark"]
ws.append(headers)
ws.row_dimensions[8].height = 28

for col in range(1, 7):
    cell = ws.cell(row=8, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# ── 5. Daily Rows 9 to 38 (30 Days) ──────────────────────────────────
for day in range(1, 31):
    row_num = 8 + day
    c_letter = col_letter(day + 1) # day 1 = col B in Jadwal Shifting
    date_str = f"{day:02d}/06/2026"

    # Dynamic formulas referencing B2 (selected consultant)
    shift_formula = f"=INDEX('Jadwal Shifting'!{c_letter}$3:{c_letter}$17, MATCH($B$2, 'Jadwal Shifting'!$A$3:$A$17, 0))"
    jam_formula = f'=IFS(B{row_num}="S1", "07:30 - 16:30", B{row_num}="S2", "15:30 - 00:30", B{row_num}="S3", "23:30 - 08:30", B{row_num}="S12", "07:30 - 00:30", B{row_num}="S23", "15:30 - 08:30", B{row_num}="OFF", "-", B{row_num}="IS", "-", TRUE, "-")'
    open_formula = f"=COUNTIFS('Open Insiden'!$C:$C, $B$2, 'Open Insiden'!$A:$A, A{row_num}&\"*\")"
    closed_formula = f"=COUNTIFS('Closed Insiden'!$C:$C, $B$2, 'Closed Insiden'!$A:$A, A{row_num}&\"*\")"
    remark_formula = f'=IFS(B{row_num}="OFF", "OFF", B{row_num}="IS", "Izin Sakit", B{row_num}="S1", "Shift 1 (Pagi)", B{row_num}="S2", "Shift 2 (Sore)", B{row_num}="S3", "Shift 3 (Malam)", B{row_num}="S12", "Shift 1 & 2", B{row_num}="S23", "Shift 2 & 3", TRUE, B{row_num})'

    ws.append([date_str, shift_formula, jam_formula, open_formula, closed_formula, remark_formula])
    ws.row_dimensions[row_num].height = 22

    row_fill = zebra_fill if day % 2 == 0 else white_fill
    for col in range(1, 7):
        cell = ws.cell(row=row_num, column=col)
        cell.font = Font(name=font_family, size=10)
        cell.fill = row_fill
        cell.border = thin_border
        if col in [1, 2, 4, 5]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Column widths
widths = {1: 15, 2: 12, 3: 20, 4: 15, 5: 15, 6: 28}
for col_idx, width in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = width

wb.save(file_path)
print("TEMPLATE_Timesheet_GoogleSheets.xlsx updated successfully with Dropdown & Premium Design!")
